import json
import datetime
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import login
from django.contrib.auth.forms import UserCreationForm
from django.contrib import messages
from django.urls import reverse
from django.utils import timezone
from django.core.paginator import Paginator
from django.db.models import Count, Q, Prefetch, Min, Max, Avg, Exists, OuterRef
from .models import Farm, RubberTree, ScanHistory, Intervention, DiseaseClass, UserSettings, get_disease_lookup
from .imaging import resize_for_storage
from . import ai_inference
from . import storage_stats
from . import direct_upload
from . import psgc


def _get_monthly_trend(request, farm=None):
    # Groups the user's actual scan history by month, counting each disease
    # type per month, for the "Reports Over Time" line chart. Keyed by
    # disease name (not a fixed field per disease) so any disease in the
    # live catalog shows up without code changes here.
    from collections import OrderedDict

    qs = ScanHistory.objects.filter(tree__farm__owner=request.user)
    if farm:
        qs = qs.filter(tree__farm=farm)
    qs = qs.select_related("tree").order_by("date")

    disease_names = [d.name for d in get_disease_lookup().values()]
    months = OrderedDict()
    for h in qs:
        key = h.date.strftime("%b %Y")
        if key not in months:
            months[key] = {"month": key, "diseases": {name: 0 for name in disease_names}}
        if h.disease in months[key]["diseases"]:
            months[key]["diseases"][h.disease] += 1

    return list(months.values())


def _get_severity_counts(request, farm=None):
    # Returns tree counts by severity tier, using database-level filtering
    # on the stored severity_score instead of iterating every tree row and
    # computing severity_label in Python.
    qs = RubberTree.objects.filter(farm=farm) if farm else RubberTree.objects.filter(farm__owner=request.user)
    return {
        "Healthy": qs.filter(disease="Healthy").count(),
        "Mild": qs.exclude(disease="Healthy").filter(severity_score__lt=34).count(),
        "Moderate": qs.exclude(disease="Healthy").filter(severity_score__gte=34, severity_score__lt=67).count(),
        "Severe": qs.exclude(disease="Healthy").filter(severity_score__gte=67).count(),
    }


# Current AI model + report template version, surfaced in the PDF report's
# metadata footer so a reader knows exactly which detector produced the
# underlying classifications and which report layout generated the file.
AI_MODEL_VERSION = "MobileNetV3-Large v2.3 (CNN Trunk Disease Classifier)"
REPORT_VERSION = "2.0"


def _get_intervention_effectiveness(request, farm=None):
    # For each intervention action type, estimates how often the treated
    # tree's CURRENT status is Healthy — a proxy for how effective that
    # action has been at resolving disease. Also returns the most recent
    # interventions logged, for a "recent activity" list.
    qs = Intervention.objects.select_related("tree", "tree__farm").filter(tree__farm__owner=request.user)
    if farm:
        qs = qs.filter(tree__farm=farm)

    by_action = {}
    for iv in qs.select_related("tree"):
        action = iv.action
        if action not in by_action:
            by_action[action] = {"action": action, "treated": 0, "recovered": 0}
        by_action[action]["treated"] += 1
        if iv.tree.disease == "Healthy":
            by_action[action]["recovered"] += 1

    effectiveness = []
    for row in by_action.values():
        row["recovery_pct"] = round(row["recovered"] / row["treated"] * 100, 1) if row["treated"] else 0
        effectiveness.append(row)
    effectiveness.sort(key=lambda r: r["treated"], reverse=True)

    recent = list(qs.order_by("-date_performed")[:8])
    return effectiveness, recent


def _get_confidence_summary(request, farm=None):
    # One-number reliability snapshot for the Key Insights list: how
    # confident the model was on average, and how many detections fell
    # low enough (<70%) that an agricultural technician should double
    # check them. Deliberately just these two numbers, not a full new
    # report section -- confidence is worth surfacing, but doesn't need
    # more room than that.
    qs = RubberTree.objects.filter(farm=farm) if farm else RubberTree.objects.filter(farm__owner=request.user)
    agg = qs.aggregate(avg=Avg("confidence"))
    avg_confidence = round(agg["avg"], 1) if agg["avg"] is not None else None
    low_threshold = 70
    low_count = qs.filter(confidence__lt=low_threshold).count()
    return {"avg": avg_confidence, "low_count": low_count, "low_threshold": low_threshold}


def _get_key_insights(total, counts, pcts, diseased, effectiveness, disease_breakdown, confidence_summary=None):
    # Builds a short list of plain-language, data-driven takeaways from
    # the current stats. Used on the reports page and in the PDF export
    # so a non-technical reader gets the "so what" up front.
    insights = []
    if total == 0:
        return ["No trees have been scanned yet — insights will appear once detection data is available."]

    insights.append(
        f"{pcts.get('Healthy', 0)}% of monitored trees ({counts.get('Healthy', 0)} of {total}) are currently Healthy."
    )

    if diseased and disease_breakdown:
        top = max(disease_breakdown, key=lambda d: d["count"])
        if top["count"] > 0:
            insights.append(
                f"{top['name']} is the most common issue detected, accounting for "
                f"{top['count']} case(s) ({top['pct']}% of all trees)."
            )
    else:
        insights.append("No active disease cases detected across the monitored trees.")

    if effectiveness:
        best = max(effectiveness, key=lambda r: (r["recovery_pct"], r["treated"]))
        if best["recovery_pct"] > 0:
            insights.append(
                f"\"{best['action']}\" is the most effective intervention on record, with "
                f"{best['recovery_pct']}% of treated trees ({best['recovered']} of {best['treated']}) now Healthy."
            )
        else:
            total_treated = sum(r["treated"] for r in effectiveness)
            insights.append(
                f"{total_treated} intervention(s) have been logged, but none of the treated trees are currently "
                f"marked Healthy yet — recovery may take longer than the current scan history covers."
            )
    else:
        insights.append("No interventions have been logged yet — recommendations below are detection-based only.")

    if confidence_summary and confidence_summary["avg"] is not None:
        c = confidence_summary
        if c["low_count"] > 0:
            insights.append(
                f"Average detection confidence is {c['avg']}%; {c['low_count']} scan(s) came in below "
                f"{c['low_threshold']}% and may be worth a manual second look."
            )
        else:
            insights.append(f"Average detection confidence is {c['avg']}%, with no low-confidence scans on record.")

    return insights[:6]


def _get_recommendations(disease_breakdown, effectiveness):
    # Builds short, actionable recommendations per disease based on what
    # share of all trees currently show it, plus a general intervention
    # recommendation drawn from the least-effective logged action (if any).
    recs = []
    for d in sorted(disease_breakdown, key=lambda r: r["pct"], reverse=True):
        if d["count"] == 0:
            continue
        if d["pct"] >= 20:
            recs.append(f"Apply treatment for {d['name']} — {d['pct']}% of all trees show symptoms ({d['count']} tree(s)); prioritize immediate intervention.")
        elif d["pct"] >= 5:
            recs.append(f"Schedule a follow-up inspection for {d['name']} — {d['count']} tree(s) affected ({d['pct']}%).")
        else:
            recs.append(f"Monitor for {d['name']} — isolated case(s) detected ({d['count']} tree(s)); recheck on the next scheduled scan.")

    if effectiveness:
        weakest = min(effectiveness, key=lambda r: r["recovery_pct"])
        if weakest["treated"] >= 2 and weakest["recovery_pct"] < 50:
            recs.append(f"Reassess the \"{weakest['action']}\" protocol — only {weakest['recovery_pct']}% of treated trees have recovered so far.")

    if not recs:
        recs.append("No disease cases at this time — maintain routine monitoring on the standard scan schedule.")

    return recs[:6]


def _get_scan_activity(request, farm=None, days=14):
    # Compares the most recent `days`-day window of scan activity against
    # the equal-length window before it, so the dashboard can show real
    # per-scan trends (volume and outcome up/down) instead of only static
    # totals. Uses the latest scan date in the data as "today" rather than
    # the server clock, so demo/seeded data (which stops at a fixed date)
    # still shows a meaningful recent window instead of an empty one.
    import datetime
    qs = ScanHistory.objects.filter(tree__farm__owner=request.user)
    if farm:
        qs = qs.filter(tree__farm=farm)

    reference_date = qs.aggregate(Max("date"))["date__max"]
    if not reference_date:
        return None

    window_start = reference_date - datetime.timedelta(days=days)
    prev_start = window_start - datetime.timedelta(days=days)

    window_qs = qs.filter(date__gt=window_start, date__lte=reference_date)
    prev_qs = qs.filter(date__gt=prev_start, date__lte=window_start)

    scans_count = window_qs.count()
    healthy_scans = window_qs.filter(disease="Healthy").count()
    diseased_scans = scans_count - healthy_scans
    trees_scanned = window_qs.values("tree").distinct().count()

    prev_count = prev_qs.count()
    prev_healthy = prev_qs.filter(disease="Healthy").count()

    health_rate = round(healthy_scans / scans_count * 100, 1) if scans_count else None
    prev_health_rate = round(prev_healthy / prev_count * 100, 1) if prev_count else None
    health_rate_delta = round(health_rate - prev_health_rate, 1) if (health_rate is not None and prev_health_rate is not None) else None
    scans_delta = scans_count - prev_count

    # Trees whose record was first created inside this window — a proxy
    # for "newly added" trees, since new trees are only ever created via
    # a fresh Disease Detection save (see disease_detection view).
    tree_qs = RubberTree.objects.filter(farm__owner=request.user)
    if farm:
        tree_qs = tree_qs.filter(farm=farm)
    new_trees = tree_qs.filter(date_scanned__gt=window_start, date_scanned__lte=reference_date).count()

    return {
        "days": days, "reference_date": reference_date, "window_start": window_start,
        "scans_count": scans_count, "scans_delta": scans_delta, "prev_scans_count": prev_count,
        "trees_scanned": trees_scanned,
        "healthy_scans": healthy_scans, "diseased_scans": diseased_scans,
        "health_rate": health_rate, "prev_health_rate": prev_health_rate, "health_rate_delta": health_rate_delta,
        "new_trees": new_trees,
    }


def _get_most_affected_farms(request, limit=8):
    # Returns farms ranked by diseased tree count, with a pre-computed bar
    # percentage relative to the worst-affected farm, for the "Most
    # Affected Areas" chart. Avoids divide-by-zero when nothing is diseased.
    ranked = []
    for f in Farm.objects.filter(owner=request.user):
        total, counts, pcts, diseased = f.get_stats()
        ranked.append({"label": f.name or f.farm_id, "farm_id": f.farm_id, "diseased": diseased, "total": total})
    ranked.sort(key=lambda r: r["diseased"], reverse=True)
    ranked = ranked[:limit]
    max_diseased = max((r["diseased"] for r in ranked), default=0)
    for r in ranked:
        r["bar_pct"] = round(r["diseased"] / max_diseased * 100, 1) if max_diseased else 0
    return ranked


def _humanize_days_ago(days):
    # Converts a day count into a short "time ago" string.
    if days <= 0:
        return "Today"
    if days == 1:
        return "1 day ago"
    if days < 30:
        return f"{days} days ago"
    months = days // 30
    return f"{months} month{'s' if months > 1 else ''} ago"


def _build_notifications(request):
    # Builds clickable notifications from the user's most recently scanned
    # diseased trees, filtered by their notification preferences (which
    # disease types, and how many days back counts as "recent") from
    # UserSettings -- see the Settings page.
    user_settings, _ = UserSettings.objects.get_or_create(user=request.user)
    enabled_diseases = user_settings.enabled_diseases()
    if not enabled_diseases:
        return []

    today = timezone.localdate()
    cutoff = today - timezone.timedelta(days=user_settings.notify_lookback_days)
    diseased_trees = (
        RubberTree.objects.select_related("farm")
        .filter(farm__owner=request.user, disease__in=enabled_diseases, date_scanned__gte=cutoff)
        .order_by("-date_scanned")[:6]
    )
    notifications = []
    disease_lookup = get_disease_lookup()
    for t in diseased_trees:
        dc = disease_lookup.get(t.disease)
        color = dc.color_hex if dc else "#22c55e"
        days_ago = (today - t.date_scanned).days
        notifications.append({
            "icon": "bi-exclamation-triangle-fill",
            "color": color,
            "msg": f"{t.disease} detected at Tree {t.tree_id}",
            "time": _humanize_days_ago(days_ago),
            "url": reverse("tree_details", args=[t.tree_id]),
        })
    return notifications


def _get_farm_or_none(request):
    # Returns the logged-in user's currently selected Farm, or None.
    farm_id = request.session.get("selected_farm_id")
    if farm_id:
        return Farm.objects.filter(pk=farm_id, owner=request.user).first()
    return None


RANGE_FILTER_DAYS = {"7": 7, "30": 30, "90": 90}


def _apply_range_filter(qs, range_param, date_field):
    # Shared "recent" filter used across the table pages: range_param is one
    # of "7"/"30"/"90" (days) or "" for all time.
    days = RANGE_FILTER_DAYS.get(range_param)
    if not days:
        return qs
    since = timezone.now().date() - datetime.timedelta(days=days)
    return qs.filter(**{f"{date_field}__gte": since})


def _severity_bounds(severity_param):
    # Mirrors RubberTree.severity_label's score thresholds, as a queryset filter.
    return {
        "Healthy": {"severity_score": 0},
        "Mild": {"severity_score__gt": 0, "severity_score__lt": 34},
        "Moderate": {"severity_score__gte": 34, "severity_score__lt": 67},
        "Severe": {"severity_score__gte": 67},
    }.get(severity_param)


def _get_trees(request, farm=None):
    # Returns the logged-in user's trees, optionally filtered to one farm.
    qs = RubberTree.objects.select_related("farm").filter(farm__owner=request.user)
    if farm:
        qs = qs.filter(farm=farm)
    return qs


def _get_stats(request, farm=None):
    # Aggregates disease counts and percentages for one farm or all the
    # user's farms, using a single GROUP BY-style query instead of looping
    # over every tree row in Python (matters once farms have thousands of
    # trees each).
    qs = RubberTree.objects.filter(farm=farm) if farm else RubberTree.objects.filter(farm__owner=request.user)
    total = qs.count()
    raw_counts = dict(qs.values_list("disease").annotate(n=Count("id")).values_list("disease", "n"))
    counts = {
        "Healthy": raw_counts.get("Healthy", 0),
        "Pink_Disease": raw_counts.get("Pink Disease", 0),
        "White_Root_Rot": raw_counts.get("White Root Rot", 0),
        "Stem_Bleeding": raw_counts.get("Stem Bleeding", 0),
    }
    pcts = {k: round(v / total * 100, 1) if total else 0 for k, v in counts.items()}
    diseased = total - counts["Healthy"]
    return total, counts, pcts, diseased


def _get_full_disease_stats(request, farm=None):
    # Every disease in the live catalog (including Healthy) with
    # count/pct/color, used to drive per-disease charts that need to
    # scale to however many disease classes currently exist.
    qs = RubberTree.objects.filter(farm=farm) if farm else RubberTree.objects.filter(farm__owner=request.user)
    total = qs.count()
    raw_counts = dict(qs.values_list("disease").annotate(n=Count("id")).values_list("disease", "n"))
    stats = []
    for d in get_disease_lookup().values():
        n = raw_counts.get(d.name, 0)
        stats.append({
            "name": d.name, "count": n,
            "pct": round(n / total * 100, 1) if total else 0,
            "color": d.color_hex, "is_healthy": d.is_healthy,
        })
    return stats


def _get_disease_breakdown(request, farm=None):
    # Same as above, minus Healthy -- for picking out e.g. the single
    # most common disease rather than just a healthy/diseased split.
    return [d for d in _get_full_disease_stats(request, farm) if not d["is_healthy"]]


def _base_context(request, farm=None):
    # Builds the context shared across all views, scoped to the logged-in user.
    all_farms = Farm.objects.filter(owner=request.user).order_by("farm_id")
    return {
        "notifications": _build_notifications(request),
        "all_farms": all_farms,
        "selected_farm": farm,
    }


def register(request):
    # Handles new user sign-up and logs them in immediately on success.
    if request.user.is_authenticated:
        return redirect("dashboard")
    if request.method == "POST":
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, "Account created. Welcome to RubberGuard!")
            return redirect("dashboard")
    else:
        form = UserCreationForm()
    return render(request, "register.html", {"form": form})


@login_required
def select_farm(request):
    # Saves the selected farm to the session and redirects back to the current page.
    farm_pk = request.POST.get("farm_pk", "")
    if farm_pk:
        if Farm.objects.filter(pk=farm_pk, owner=request.user).exists():
            request.session["selected_farm_id"] = int(farm_pk)
    else:
        request.session.pop("selected_farm_id", None)
    next_url = request.POST.get("next", "/")
    return redirect(next_url)


@login_required
def settings_view(request):
    # Account info, password change, and notification preferences, each
    # submitted as its own form (distinguished by the "form_name" field)
    # so saving one section never touches the others.
    from django.contrib.auth.forms import PasswordChangeForm
    from django.contrib.auth import update_session_auth_hash

    user_settings, _ = UserSettings.objects.get_or_create(user=request.user)
    password_form = PasswordChangeForm(request.user)

    if request.method == "POST":
        form_name = request.POST.get("form_name")

        if form_name == "account":
            request.user.email = request.POST.get("email", "").strip()
            request.user.first_name = request.POST.get("first_name", "").strip()
            request.user.last_name = request.POST.get("last_name", "").strip()
            request.user.save()
            messages.success(request, "Account info updated.")
            return redirect("settings")

        elif form_name == "password":
            password_form = PasswordChangeForm(request.user, request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)  # keeps the user logged in
                messages.success(request, "Password changed.")
                return redirect("settings")
            messages.error(request, "Please fix the errors below.")

        elif form_name == "notifications":
            try:
                user_settings.notify_lookback_days = max(1, int(request.POST.get("notify_lookback_days", 7)))
            except ValueError:
                user_settings.notify_lookback_days = 7
            # Checked box = notify (matches settings.html rendering below),
            # so anything NOT posted for a known disease name is muted.
            checked = set(request.POST.getlist("notify_disease"))
            user_settings.notify_muted_diseases = [
                d.name for d in get_disease_lookup().values()
                if not d.is_healthy and d.name not in checked
            ]
            user_settings.save()
            messages.success(request, "Notification preferences updated.")
            return redirect("settings")

    ctx = _base_context(request)
    ctx.update({
        "page": "settings",
        "password_form": password_form,
        "user_settings": user_settings,
        "notifiable_diseases": [d for d in get_disease_lookup().values() if not d.is_healthy],
    })
    return render(request, "settings.html", ctx)


@login_required
def farm_list(request):
    # Displays a list of all farms owned by the logged-in user.
    # tree_count/avg_lat/avg_lng are annotated here (one GROUP BY query for
    # the whole page) rather than calling farm.trees.count() / get_center()
    # per row in the template, which would re-query per farm card.
    farms_qs = Farm.objects.filter(owner=request.user).annotate(
        tree_count=Count("trees"),
        avg_lat=Avg("trees__lat"),
        avg_lng=Avg("trees__lng"),
    ).order_by("farm_id")
    search_q = request.GET.get("q", "").strip()
    if search_q:
        farms_qs = farms_qs.filter(
            Q(farm_id__icontains=search_q) | Q(name__icontains=search_q)
            | Q(owner_name__icontains=search_q) | Q(location__icontains=search_q)
        )
    paginator = Paginator(farms_qs, 25)
    page_number = request.GET.get("page", 1)
    farms = paginator.get_page(page_number)
    ctx = _base_context(request)
    ctx.update({"page": "farm_list", "farms": farms, "search_q": search_q, "psgc_regions": psgc.get_regions()})
    return render(request, "farm_list.html", ctx)


@login_required
def psgc_provinces(request, region_code):
    # Powers the Province dropdown once a Region is picked on Add Farm.
    return JsonResponse({"results": psgc.get_provinces(region_code)})


@login_required
def psgc_cities(request, province_code):
    # Powers the City/Municipality dropdown once a Province is picked.
    return JsonResponse({"results": psgc.get_cities(province_code)})


@login_required
def psgc_barangays(request, city_code):
    # Powers the Barangay dropdown once a City is picked. This is the only
    # one of the 3 that touches the 42k-row barangay dataset, and only for
    # the one city asked about -- never the full list.
    return JsonResponse({"results": psgc.get_barangays(city_code)})


@login_required
def farm_create(request):
    # Handles the Add Farm form, creating a new farm owned by the logged-in user.
    if request.method == "POST":
        farm_id = request.POST.get("farm_id", "").strip()
        name = request.POST.get("name", "").strip()
        owner_name = request.POST.get("owner_name", "").strip()
        location = request.POST.get("location", "").strip()
        # Posted as plain names (not PSGC codes) by the cascading selects
        # in farm_list.html -- see wireAddressCascade() in that template.
        region = request.POST.get("region_name", "").strip()
        province = request.POST.get("province_name", "").strip()
        city_municipality = request.POST.get("city_name", "").strip()
        barangay = request.POST.get("barangay_name", "").strip()

        if not farm_id or not name or not owner_name:
            messages.error(request, "Farm ID, name, and owner name are required.")
            return redirect("farm_list")

        if Farm.objects.filter(owner=request.user, farm_id=farm_id).exists():
            messages.error(request, f"You already have a farm with ID '{farm_id}'.")
            return redirect("farm_list")

        # No center/boundary here on purpose -- Farm.center_lat/center_lng
        # just take their model defaults (a general Zamboanga City point) as
        # a placeholder until this farm's first tree scan comes in, at which
        # point get_center()/get_boundary_polygon() take over and compute
        # the real thing from actual tree GPS data instead.
        Farm.objects.create(
            owner=request.user,
            farm_id=farm_id,
            name=name,
            owner_name=owner_name,
            region=region,
            province=province,
            city_municipality=city_municipality,
            barangay=barangay,
            location=location,
        )
        messages.success(request, f"Farm '{name}' added successfully.")
        return redirect("farm_list")

    return redirect("farm_list")


@login_required
def farm_detail(request, farm_id):
    # Displays details, stats, and trees for a single farm owned by the user.
    farm = get_object_or_404(Farm, farm_id=farm_id, owner=request.user)
    total, counts, pcts, diseased = farm.get_stats()
    disease_stats = _get_full_disease_stats(request, farm)
    trees_qs = farm.trees.all().order_by("tree_id")
    paginator = Paginator(trees_qs, 25)
    page_number = request.GET.get("page", 1)
    trees_page = paginator.get_page(page_number)
    ctx = _base_context(request)
    ctx.update({
        "page": "farm_list",
        "farm": farm,
        "trees": trees_page,
        "total": total, "counts": counts, "pcts": pcts, "diseased": diseased,
        "disease_stats": disease_stats,
    })
    return render(request, "farm_detail.html", ctx)


def ping(request):
    # Deliberately outside @login_required and does zero DB/template work --
    # this only exists as a cheap target for the offline-banner's real
    # connectivity probe (see offline_queue.js), which may poll every ~15s.
    from django.http import HttpResponse
    return HttpResponse(status=204)


@login_required
def dashboard(request):
    # Renders the dashboard: summary cards, recent detections, and a map preview.
    farm = _get_farm_or_none(request)
    total, counts, pcts, diseased = _get_stats(request, farm)
    severity_counts = _get_severity_counts(request, farm)
    # Bar-fill percentages for the Severity Breakdown panel, relative to
    # the largest tier so the bars are visually comparable (same approach
    # _get_most_affected_farms uses for its bar_pct field).
    severity_max = max(severity_counts.values()) or 1
    severity_bar_pcts = {k: round(v / severity_max * 100, 1) for k, v in severity_counts.items()}
    confidence_summary = _get_confidence_summary(request, farm)
    trees = list(
        _get_trees(request, farm).select_related("farm")
        .prefetch_related(
            Prefetch("history", queryset=ScanHistory.objects.order_by("-date")),
            Prefetch("interventions", queryset=Intervention.objects.order_by("-date_performed")),
        )
        .order_by("-date_scanned")[:6]
    )
    recent = [t.to_dict() for t in trees]

    farm_count = Farm.objects.filter(owner=request.user).count()
    scan_activity = _get_scan_activity(request, farm)
    effectiveness, recent_interventions = _get_intervention_effectiveness(request, farm)
    top_interventions = [e for e in effectiveness if e["treated"] > 0][:3]
    disease_stats = _get_full_disease_stats(request, farm)

    ctx = _base_context(request, farm)
    ctx.update({
        "page": "dashboard",
        "total": total, "counts": counts, "pcts": pcts, "diseased": diseased,
        "severity_counts": severity_counts,
        "severity_bar_pcts": severity_bar_pcts,
        "confidence_summary": confidence_summary,
        "healthy_count": counts["Healthy"],
        "areas_with_cases": Farm.objects.filter(
            owner=request.user, trees__disease__isnull=False
        ).exclude(trees__disease="Healthy").distinct().count() if not farm else (1 if diseased else 0),
        "farm_count": farm_count,
        "recent": recent,
        "scan_activity": scan_activity,
        "recent_interventions": recent_interventions[:6],
        "top_interventions": top_interventions,
        "disease_stats": disease_stats,
        "latest_scan": recent[0]["date_scanned"] if recent else "—",
    })
    return render(request, "dashboard.html", ctx)


def _farm_map_bounds(farm):
    # Returns a bounding box around a farm's actual trees (falling back to
    # its boundary radius if it has no trees yet), used to zoom-lock any
    # map so the user can zoom in freely but not zoom out past their own
    # farm's extent. Shared by farm_map and the reports page map.
    bounds = farm.trees.aggregate(
        min_lat=Min("lat"), max_lat=Max("lat"),
        min_lng=Min("lng"), max_lng=Max("lng"),
    )
    if bounds["min_lat"] is None:
        center_lat, center_lng = farm.get_center()
        deg_pad = max(farm.boundary_radius_m, 200) / 111000
        bounds = {
            "min_lat": center_lat - deg_pad, "max_lat": center_lat + deg_pad,
            "min_lng": center_lng - deg_pad, "max_lng": center_lng + deg_pad,
        }
    else:
        # Guard against a zero-height/width box (e.g. every tree shares
        # the same lat or lng so far) -- Leaflet's fitBounds/setMaxBounds
        # on a truly zero-area box can zoom to an unusable extreme. Must
        # match get_boundary_polygon()'s own fallback size (not a small
        # fixed value): otherwise the map fits/zooms tighter than that
        # square, pushing its edges off-screen so the boundary never
        # actually looks like it's there.
        min_pad = max(farm.boundary_radius_m, 200) / 111000
        if bounds["max_lat"] - bounds["min_lat"] < min_pad:
            bounds["min_lat"] -= min_pad
            bounds["max_lat"] += min_pad
        if bounds["max_lng"] - bounds["min_lng"] < min_pad:
            bounds["min_lng"] -= min_pad
            bounds["max_lng"] += min_pad
    return bounds


@login_required
def farm_map(request):
    # Renders the interactive Leaflet map for exactly one farm at a time,
    # with its own page-scoped farm selector (a ?farm=<id> query param)
    # independent of the sidebar's "Active Farm" selector. This avoids the
    # sidebar showing a specific farm as "selected" just because the map
    # needed to default to one — the sidebar keeps reflecting the user's
    # actual session-wide choice (which may genuinely be "All Farms").
    sidebar_farm = _get_farm_or_none(request)
    all_user_farms = Farm.objects.filter(owner=request.user).order_by("farm_id")

    map_farm_id = request.GET.get("farm", "")
    if map_farm_id:
        map_farm = all_user_farms.filter(farm_id=map_farm_id).first()
    else:
        map_farm = sidebar_farm or all_user_farms.first()

    if not map_farm:
        ctx = _base_context(request, sidebar_farm)
        ctx.update({"page": "farm_map", "no_farms": True})
        return render(request, "farm_map.html", ctx)

    total, counts, pcts, diseased = map_farm.get_stats()
    trees_qs = map_farm.trees.annotate(
        has_intervention=Exists(Intervention.objects.filter(tree=OuterRef("pk")))
    )
    markers_json = json.dumps([t.to_marker_dict() for t in trees_qs])
    bounds = _farm_map_bounds(map_farm)
    boundary_polygon = json.dumps(map_farm.get_boundary_polygon())

    # Built fresh from the live DiseaseClass catalog (not the fixed-key
    # counts dict above) so the filter dropdown, legend, and top-disease
    # card all pick up newly added disease classes automatically -- same
    # reasoning as RubberTree.color/disease_key.
    raw_counts = dict(
        map_farm.trees.values_list("disease").annotate(n=Count("id")).values_list("disease", "n")
    )
    disease_stats = []
    top_disease = None
    for d in DiseaseClass.objects.all():
        n = raw_counts.get(d.name, 0)
        disease_stats.append({"name": d.name, "count": n, "color": d.color_hex, "is_healthy": d.is_healthy})
        if not d.is_healthy and (top_disease is None or n > top_disease["count"]):
            top_disease = {"name": d.name, "count": n, "color": d.color_hex}
    disease_colors = {d["name"]: d["color"] for d in disease_stats}

    # Uses the sidebar's own selected_farm (not map_farm) for shared
    # context, so the sidebar dropdown never appears to change just from
    # visiting this page.
    ctx = _base_context(request, sidebar_farm)
    ctx.update({
        "page": "farm_map",
        "map_farm": map_farm,
        "all_user_farms": all_user_farms,
        "markers_json": markers_json,
        "map_bounds": json.dumps(bounds),
        "map_farm_boundary": boundary_polygon,
        "total": total, "counts": counts, "diseased": diseased,
        "disease_stats": disease_stats,
        "top_disease": top_disease,
        "disease_colors_json": json.dumps(disease_colors),
    })
    return render(request, "farm_map.html", ctx)


@login_required
def tree_marker_detail(request, tree_id):
    # Returns full marker popup detail (recommended action, notes, latest
    # inspector, latest intervention) for a single tree as JSON. Called via
    # AJAX only when a marker is actually clicked, instead of embedding
    # this for every tree on initial map load.
    from django.http import JsonResponse, Http404
    tree = (
        RubberTree.objects.select_related("farm")
        .prefetch_related(
            Prefetch("history", queryset=ScanHistory.objects.order_by("-date")),
            Prefetch("interventions", queryset=Intervention.objects.order_by("-date_performed")),
        )
        .filter(tree_id=tree_id, farm__owner=request.user)
        .first()
    )
    if not tree:
        raise Http404("Tree not found")
    return JsonResponse(tree.to_dict())


@login_required
def service_worker_js(request):
    # Served from /detection/sw.js, not /static/js/sw.js -- see the
    # comment at the top of static/js/sw.js for why that path matters.
    from django.http import HttpResponse
    from pathlib import Path
    sw_path = Path(__file__).resolve().parent.parent / "static" / "js" / "sw.js"
    return HttpResponse(sw_path.read_text(), content_type="application/javascript")

@login_required
def disease_detection(request):
    # Renders the disease detection upload page.
    farm = _get_farm_or_none(request)
    ctx = _base_context(request, farm)
    ctx.update({
        "page": "disease_detection",
        "disease_classes": DiseaseClass.objects.all().order_by("display_order", "name"),
        # Analyze Images is disabled in the template when this is False,
        # rather than the page silently making up a result.
        "ai_enabled": ai_inference.AI_MODEL_ENABLED,
    })

    # "Scan Again" from a tree's detail page arrives here with these two,
    # so the farm/tree fields come pre-filled instead of the farmer having
    # to retype the tree's exact code from memory (see tree_details.html).
    # rescan_farm_pk is compared as an int against farm.pk in the template,
    # not just forwarded as the raw GET string, so a tampered/bad value
    # just fails to match any option instead of ever being trusted.
    rescan_tree_id = request.GET.get("rescan_tree_id", "").strip()
    try:
        rescan_farm_pk = int(request.GET.get("rescan_farm_pk", ""))
    except ValueError:
        rescan_farm_pk = None
    if rescan_tree_id:
        ctx.update({"rescan_tree_id": rescan_tree_id, "rescan_farm_pk": rescan_farm_pk})

    return render(request, "disease_detection.html", ctx)


@login_required
def bulk_upload(request):
    # Renders the bulk upload page. All the real work (clustering photos
    # into trees by timestamp/GPS, letting the user confirm/fix groups,
    # then analyzing+saving each) happens client-side in bulk_detection.js,
    # reusing the same /detection/analyze/ and /detection/save/ endpoints
    # the single-tree flow and offline sync already use -- one call per
    # confirmed tree, nothing new server-side.
    farm = _get_farm_or_none(request)
    ctx = _base_context(request, farm)
    # Each farm's live center, for defaulting the "set location manually"
    # map when a group has no GPS of its own.
    farm_centers = {str(f.pk): list(f.get_center()) for f in ctx["all_farms"]}
    ctx.update({
        "page": "bulk_upload",
        "ai_enabled": ai_inference.AI_MODEL_ENABLED,
        "farm_centers": farm_centers,
    })
    return render(request, "bulk_upload.html", ctx)


def _severity_label_for(disease, confidence):
    # Mirrors RubberTree.severity_label's thresholds (models.py), for a
    # result that hasn't been saved as a RubberTree yet -- see
    # analyze_detection() below.
    disease_class = get_disease_lookup().get(disease)
    if (disease_class and disease_class.is_healthy) or not confidence:
        return "Healthy"
    if confidence < 34:
        return "Mild"
    if confidence < 67:
        return "Moderate"
    return "Severe"


@login_required
def analyze_detection(request):
    # Runs the configured AI model on the two captured photos and returns
    # its result as JSON. Called by the Analyze Images button so what the
    # user reviews before saving is the real model output. save_detection()
    # independently re-runs this at save time too -- that copy is the one
    # actually trusted for what gets stored, since these hidden form
    # fields could still be edited in devtools before submit.
    from django.http import JsonResponse

    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=405)

    if not ai_inference.AI_MODEL_ENABLED:
        return JsonResponse(
            {"error": "AI model isn't configured. Set AI_MODEL_ENDPOINT_URL to enable detection."},
            status=503,
        )

    root_file = request.FILES.get("root_image")
    trunk_file = request.FILES.get("trunk_image")
    if not root_file or not trunk_file:
        return JsonResponse({"error": "Both root and trunk images are required."}, status=400)

    try:
        result = ai_inference.classify_images(root_file.read(), trunk_file.read())
    except ai_inference.InferenceError as exc:
        return JsonResponse({"error": str(exc)}, status=502)

    disease_class = get_disease_lookup().get(result["disease"])
    severity_label = _severity_label_for(result["disease"], result["confidence"])
    result["action"] = disease_class.recommendation_for(severity_label) if disease_class else ""
    return JsonResponse(result)


@login_required
def request_upload_url(request):
    # Issues a short-lived presigned R2 PUT URL for one scan photo.
    # Called twice by the frontend (once for the root photo, once for the
    # trunk) right before "Save Result" -- see static/js/upload_direct.js.
    # Frontend has already compressed the photo to WebP itself at this
    # point, so all this endpoint does is hand back somewhere to put it.
    from django.http import JsonResponse

    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=405)

    kind = request.POST.get("kind")
    if kind not in ("roots", "trunks"):
        return JsonResponse({"error": "kind must be 'roots' or 'trunks'"}, status=400)

    try:
        data = direct_upload.generate_upload_url(kind)
    except direct_upload.DirectUploadUnavailable as exc:
        # Cloud storage isn't configured yet -- tell the client so it can
        # fall back to the normal multipart file upload instead.
        return JsonResponse({"error": str(exc), "fallback": True}, status=503)

    return JsonResponse(data)


@login_required
def save_detection(request):
    # Saves a CNN detection result. Creates a
    # new tree, or — if tree_id matches an existing tree on this farm —
    # appends a new ScanHistory entry and updates that tree's current
    # state instead. The latter is what makes progression tracking
    # (comment 9) reachable through real usage, not just seeded demo data.
    import datetime
    from django.http import JsonResponse

    if request.method != "POST":
        return redirect("disease_detection")

    # The offline sync queue (offline_queue.js) replays this same endpoint
    # via fetch() instead of a real form submission, since there's no page
    # to redirect. It sends this header to ask for JSON back instead.
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

    def fail(message, status=400):
        if is_ajax:
            return JsonResponse({"error": message}, status=status)
        messages.error(request, message)
        return redirect("disease_detection")

    farm_pk = request.POST.get("farm_pk")
    disease = request.POST.get("disease", "Healthy")
    confidence = request.POST.get("confidence", "0")
    tree_id = request.POST.get("tree_id", "").strip()
    root_condition = request.POST.get("root_condition", "").strip()

    # GPS is required, not just preferred. Client-side (disease_detection.js),
    # a fresh camera capture can fall back to a live device position when
    # the photo has no EXIF GPS, but a browsed/dropped file can't -- using
    # the phone's current position for an existing photo would mislabel
    # where it was actually taken, so that path is rejected outright with
    # no location instead. Either way, this should basically always be
    # present by the time a request gets here, but it's still re-checked
    # server-side rather than trusted, same reasoning as the disease/
    # confidence re-check below.
    try:
        tree_lat = float(request.POST.get("lat") or "")
        tree_lng = float(request.POST.get("lng") or "")
    except ValueError:
        tree_lat = tree_lng = None
    if tree_lat is None or tree_lng is None:
        return fail("This scan is missing a GPS location and can't be saved. Retake the photo somewhere the camera/GPS can get a location.")

    captured_at = None
    captured_at_raw = request.POST.get("captured_at", "").strip()
    if captured_at_raw:
        try:
            captured_at = datetime.datetime.fromisoformat(captured_at_raw.replace("Z", "+00:00"))
        except ValueError:
            pass  # bad/missing timestamp isn't worth failing the save over

    try:
        farm = Farm.objects.get(pk=farm_pk, owner=request.user)
    except (Farm.DoesNotExist, ValueError, TypeError):
        return fail("This scan has no valid farm attached and can't be saved.")

    root_image_resized = None
    trunk_image_resized = None
    root_file_bytes = None
    trunk_file_bytes = None

    # Preferred path: the frontend already compressed the photo to WebP
    # and uploaded it straight to R2 (see upload_direct.js), and just
    # tells us the object key here. No re-upload -- assigning the key
    # string directly to the ImageField marks it as an existing file
    # rather than a new one to save.
    root_image_key = request.POST.get("root_image_key", "").strip()
    trunk_image_key = request.POST.get("trunk_image_key", "").strip()
    if root_image_key:
        root_file_bytes = direct_upload.fetch_uploaded_bytes(root_image_key)
        root_image_resized = root_image_key
    if trunk_image_key:
        trunk_file_bytes = direct_upload.fetch_uploaded_bytes(trunk_image_key)
        trunk_image_resized = trunk_image_key

    # Fallback path: a raw file came through the multipart form instead
    # (older browsers without canvas/WebP support, direct uploads
    # unavailable because cloud storage isn't configured, or this is an
    # offline-queued scan being synced -- those always use this path,
    # since they were never uploaded to R2 while offline).
    if not root_image_key and request.FILES.get("root_image"):
        f = request.FILES["root_image"]
        root_file_bytes = f.read()
        f.seek(0)
        root_image_resized = resize_for_storage(f, f.name)
    if not trunk_image_key and request.FILES.get("trunk_image"):
        f = request.FILES["trunk_image"]
        trunk_file_bytes = f.read()
        f.seek(0)
        trunk_image_resized = resize_for_storage(f, f.name)

    # Re-runs inference server-side rather than trusting the hidden
    # save-form fields outright -- they already hold the real result from
    # analyze_detection() above, but could still be edited in devtools
    # before submit. Falls back to those submitted values only if this
    # second call fails transiently, so a flaky or cold-starting model
    # host never blocks saving a scan outright.
    if ai_inference.AI_MODEL_ENABLED and root_file_bytes and trunk_file_bytes:
        try:
            result = ai_inference.classify_images(root_file_bytes, trunk_file_bytes)
            disease = result["disease"]
            confidence = result["confidence"]
            if result.get("root_condition"):
                root_condition = result["root_condition"]
        except ai_inference.InferenceError:
            pass

    if not tree_id:
        existing = farm.trees.count()
        tree_id = f"{farm.farm_id}-RT-{existing + 1:04d}"
    elif not tree_id.startswith(f"{farm.farm_id}-"):
        # Ensure user-typed tree IDs stay globally unique by always
        # prefixing with the farm ID, since tree_id is only unique per farm
        # at the database level but URLs/lookups treat it as the sole key.
        tree_id = f"{farm.farm_id}-{tree_id}"

    try:
        existing_tree = RubberTree.objects.filter(tree_id=tree_id).first()

        if existing_tree and existing_tree.farm_id != farm.id:
            return fail(f"Tree ID '{tree_id}' already exists on a different farm.")

        if existing_tree:
            # Rescan: update the tree's current snapshot and blank out
            # recommended_action so save() re-derives it for the new
            # disease/severity combination.
            tree = existing_tree
            tree.disease = disease
            tree.confidence = float(confidence)
            tree.root_condition = root_condition
            tree.severity_score = 0.0
            tree.recommended_action = ""
            tree.date_scanned = datetime.date.today()
            tree.captured_at = captured_at
            tree.lat = tree_lat
            tree.lng = tree_lng
            if root_image_resized:
                tree.root_image = root_image_resized
            if trunk_image_resized:
                tree.trunk_image = trunk_image_resized
            tree.save()
        else:
            tree = RubberTree.objects.create(
                farm=farm, tree_id=tree_id,
                lat=tree_lat, lng=tree_lng,
                disease=disease, confidence=float(confidence),
                root_condition=root_condition,
                date_scanned=datetime.date.today(), captured_at=captured_at,
                root_image=root_image_resized, trunk_image=trunk_image_resized,
                # recommended_action is left blank here so RubberTree.save() derives
                # it from disease + severity_label (see SEVERITY_RECOMMENDATIONS).
            )

        ScanHistory.objects.create(
            tree=tree, date=datetime.date.today(), captured_at=captured_at,
            disease=disease, confidence=float(confidence), inspector=request.user.username,
            root_condition=root_condition,
            root_image=root_image_resized, trunk_image=trunk_image_resized,
        )
    except Exception as exc:
        # Without this, any unexpected DB/storage error here (e.g. a
        # pending migration not yet applied) returns Django's HTML error
        # page instead of JSON -- offline_queue.js can't parse that for a
        # real reason and falls back to a generic "Save failed", which is
        # useless for actually debugging what happened.
        return fail(f"Save failed: {exc}", status=500)

    verb = "updated" if existing_tree else "saved"
    messages.success(request, f"Detection {verb} for tree '{tree_id}'.")

    if is_ajax:
        return JsonResponse({"ok": True, "tree_id": tree_id, "verb": verb})
    return redirect("tree_details", tree_id=tree_id)


@login_required
def tree_inventory(request):
    # Renders the tree inventory table, filtered by the selected farm if
    # set, plus optional search/disease/farm query params, and paginated
    # so large datasets don't try to render thousands of rows at once.
    farm = _get_farm_or_none(request)
    total, counts, pcts, diseased = _get_stats(request, farm)
    disease_breakdown = _get_disease_breakdown(request, farm)
    top_disease = max(disease_breakdown, key=lambda d: d["count"], default=None)
    if top_disease and top_disease["count"] == 0:
        top_disease = None
    trees_qs = _get_trees(request, farm).select_related("farm").order_by("tree_id")

    search_q = request.GET.get("q", "").strip()
    disease_filter = request.GET.get("disease", "").strip()
    farm_filter = request.GET.get("farm", "").strip()
    range_filter = request.GET.get("range", "").strip()
    severity_filter = request.GET.get("severity", "").strip()

    if search_q:
        trees_qs = trees_qs.filter(tree_id__icontains=search_q)
    if disease_filter:
        trees_qs = trees_qs.filter(disease=disease_filter)
    if farm_filter:
        trees_qs = trees_qs.filter(farm__farm_id=farm_filter)
    trees_qs = _apply_range_filter(trees_qs, range_filter, "date_scanned")
    severity_bounds = _severity_bounds(severity_filter)
    if severity_bounds:
        trees_qs = trees_qs.filter(**severity_bounds)

    paginator = Paginator(trees_qs, 25)
    page_number = request.GET.get("page", 1)
    trees_page = paginator.get_page(page_number)

    ctx = _base_context(request, farm)
    ctx.update({
        "page": "tree_inventory",
        "trees": trees_page,
        "total": total, "counts": counts, "diseased": diseased,
        "top_disease": top_disease,
        "search_q": search_q, "disease_filter": disease_filter, "farm_filter": farm_filter,
        "range_filter": range_filter, "severity_filter": severity_filter,
        "disease_classes": DiseaseClass.objects.all().order_by("display_order", "name"),
    })
    return render(request, "tree_inventory.html", ctx)


@login_required
def tree_details(request, tree_id):
    # Renders the detail page for a single tree, including its scan history.
    from django.http import Http404
    tree = (
        RubberTree.objects.select_related("farm")
        .filter(tree_id=tree_id, farm__owner=request.user)
        .first()
    )
    if not tree:
        raise Http404("Tree not found")
    history_qs = tree.history.all()
    paginator = Paginator(history_qs, 25)
    page_number = request.GET.get("page", 1)
    history_page = paginator.get_page(page_number)
    farm = _get_farm_or_none(request)
    ctx = _base_context(request, farm)
    ctx.update({
        "page": "tree_inventory",
        "tree": tree,
        "history": history_page,
        # Chart needs every scan to plot an accurate trend, not just the
        # current table page -- kept separate from the paginated "history"
        # context var used for the table below.
        "history_chart_json": json.dumps(list(history_qs.values("date", "confidence", "disease")), default=str),
        "progression": tree.get_progression_trend(),
        "tree_farm_boundary": json.dumps(tree.farm.get_boundary_polygon()),
        # Legend + chart colors, pulled from the DiseaseClass catalog.
        "legend_diseases": DiseaseClass.objects.all().order_by("display_order", "name"),
        "disease_colors_json": json.dumps({name: dc.color_hex for name, dc in get_disease_lookup().items()}),
    })
    return render(request, "tree_details.html", ctx)


@login_required
def reports(request):
    # Renders the reports page: severity distribution, trend over time,
    # most-affected areas, detection summary, an interactive heatmap, and
    # a per-farm breakdown table.
    farm = _get_farm_or_none(request)
    total, counts, pcts, diseased = _get_stats(request, farm)
    severity_counts = _get_severity_counts(request, farm)
    monthly = _get_monthly_trend(request, farm)
    most_affected = _get_most_affected_farms(request)
    effectiveness, recent_interventions = _get_intervention_effectiveness(request, farm)
    disease_breakdown = _get_disease_breakdown(request, farm)
    disease_stats = _get_full_disease_stats(request, farm)
    confidence_summary = _get_confidence_summary(request, farm)
    key_insights = _get_key_insights(total, counts, pcts, diseased, effectiveness, disease_breakdown, confidence_summary)
    recommendations = _get_recommendations(disease_breakdown, effectiveness)

    farm_summaries = []
    for f in Farm.objects.filter(owner=request.user).order_by("farm_id"):
        ft, fc, fp, fd = f.get_stats()
        farm_summaries.append({
            "farm": f, "total": ft, "counts": fc, "pcts": fp, "diseased": fd,
        })

    ctx = _base_context(request, farm)
    ctx.update({
        "page": "reports",
        "total": total, "counts": counts, "pcts": pcts, "diseased": diseased,
        "disease_stats": disease_stats,
        "severity_counts": severity_counts,
        "monthly": monthly,
        "most_affected": most_affected,
        "farm_summaries": farm_summaries,
        "effectiveness": effectiveness,
        "recent_interventions": recent_interventions,
        "key_insights": key_insights,
        "recommendations": recommendations,
        "ai_model_version": AI_MODEL_VERSION,
        "report_version": REPORT_VERSION,
        "generated_at": timezone.now(),
    })
    return render(request, "reports.html", ctx)


@login_required
def interventions_map(request):
    # Renders a dedicated map showing every tree that has had at least one
    # intervention logged, so users can see where treatment work has
    # actually happened rather than just where disease was detected. This
    # set is naturally small (only treated trees), so the fuller to_dict()
    # with intervention/inspector detail is fine here.
    farm = _get_farm_or_none(request)
    trees_qs = (
        _get_trees(request, farm).select_related("farm")
        .prefetch_related(
            Prefetch("history", queryset=ScanHistory.objects.order_by("-date")),
            Prefetch("interventions", queryset=Intervention.objects.order_by("-date_performed")),
        )
        .filter(interventions__isnull=False).distinct()
    )
    ctx = _base_context(request, farm)
    ctx.update({
        "page": "interventions",
        "trees_json": json.dumps([t.to_dict() for t in trees_qs]),
        "intervention_count": Intervention.objects.filter(tree__farm__owner=request.user).count(),
    })
    return render(request, "interventions_map.html", ctx)


@login_required
def interventions_log(request):
    # Lists all logged interventions for the user's trees, most recent
    # first, paginated so a farm with many interventions doesn't render
    # them all on one page.
    farm = _get_farm_or_none(request)
    qs = Intervention.objects.select_related("tree", "tree__farm", "performed_by").filter(
        tree__farm__owner=request.user
    )
    if farm:
        qs = qs.filter(tree__farm=farm)

    search_q = request.GET.get("q", "").strip()
    action_filter = request.GET.get("action", "").strip()
    farm_filter = request.GET.get("farm", "").strip()
    range_filter = request.GET.get("range", "").strip()

    if search_q:
        qs = qs.filter(tree__tree_id__icontains=search_q)
    if action_filter:
        qs = qs.filter(action=action_filter)
    if farm_filter:
        qs = qs.filter(tree__farm__farm_id=farm_filter)
    qs = _apply_range_filter(qs, range_filter, "date_performed")

    qs = qs.order_by("-date_performed", "-created_at")

    paginator = Paginator(qs, 25)
    page_number = request.GET.get("page", 1)
    interventions_page = paginator.get_page(page_number)

    # Pre-built {farm_pk: [{tree_id, disease}, ...]} lookup for the
    # checklist-mode tree selector in the log-intervention form. Capped
    # since farms can have thousands of trees; range mode (the primary
    # selection method) has no such limit.
    CHECKLIST_TREE_LIMIT = 300
    farm_trees = {}
    for f in Farm.objects.filter(owner=request.user):
        farm_trees[str(f.pk)] = [
            {"tree_id": t.tree_id, "disease": t.disease}
            for t in f.trees.order_by("tree_id")[:CHECKLIST_TREE_LIMIT]
        ]

    ctx = _base_context(request, farm)
    ctx.update({
        "page": "interventions",
        "interventions": interventions_page,
        "farm_trees_json": json.dumps(farm_trees),
        "search_q": search_q, "action_filter": action_filter,
        "farm_filter": farm_filter, "range_filter": range_filter,
        "action_choices": Intervention.ACTION_CHOICES,
    })
    return render(request, "interventions_log.html", ctx)


@login_required
def intervention_create(request):
    # Logs a new intervention against one or more trees, selected either by
    # a tree ID range (e.g. RT-001 to RT-005) or by explicit tree IDs
    # (e.g. checked off via the map).
    import datetime

    if request.method != "POST":
        return redirect("interventions_log")

    action = request.POST.get("action", "Other")
    date_performed = request.POST.get("date_performed") or str(datetime.date.today())
    notes = request.POST.get("notes", "").strip()
    farm_pk = request.POST.get("farm_pk")
    selection_mode = request.POST.get("selection_mode", "single")

    # Guard against farm_pk being missing/blank -- filtering pk="" crashes
    # with a ValueError instead of a clean 404 (every other farm_pk lookup
    # in this file already has an equivalent guard; this one didn't).
    if not farm_pk:
        messages.error(request, "No farm was selected for this intervention.")
        return redirect("interventions_log")

    farm = get_object_or_404(Farm, pk=farm_pk, owner=request.user)
    trees = RubberTree.objects.none()

    if selection_mode == "range":
        start_id = request.POST.get("range_start", "").strip()
        end_id = request.POST.get("range_end", "").strip()
        all_ids = list(farm.trees.order_by("tree_id").values_list("tree_id", flat=True))
        try:
            start_i = all_ids.index(start_id)
            end_i = all_ids.index(end_id)
            selected_ids = all_ids[min(start_i, end_i):max(start_i, end_i) + 1]
            trees = farm.trees.filter(tree_id__in=selected_ids)
        except ValueError:
            messages.error(request, "Invalid tree ID range. Check that both IDs exist on this farm.")
            return redirect("interventions_log")
    else:
        tree_ids = request.POST.getlist("tree_ids")
        trees = farm.trees.filter(tree_id__in=tree_ids)

    if not trees.exists():
        messages.error(request, "No trees were selected for this intervention.")
        return redirect("interventions_log")

    created = 0
    for tree in trees:
        Intervention.objects.create(
            tree=tree, performed_by=request.user, action=action,
            date_performed=date_performed, notes=notes,
        )
        created += 1

    messages.success(request, f"Logged '{action}' on {created} tree(s).")
    return redirect("interventions_log")


def _export_rows(request):
    # Returns the user's trees and a filename-safe label for the current export.
    farm = _get_farm_or_none(request)
    trees = _get_trees(request, farm).select_related("farm").order_by("farm__farm_id", "tree_id")
    label = farm.farm_id if farm else "all_farms"
    return trees, label


@login_required
def export_csv(request):
    # Exports a summary-only CSV: overall disease totals and a per-farm breakdown.
    import csv
    from django.http import HttpResponse

    farm = _get_farm_or_none(request)
    total, counts, pcts, diseased = _get_stats(request, farm)
    disease_stats = _get_full_disease_stats(request, farm)
    label = farm.farm_id if farm else "all_farms"

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="rubberguard_summary_{label}.csv"'

    writer = csv.writer(response)
    writer.writerow(["RubberGuard Disease Detection Summary"])
    writer.writerow(["Scope", farm.name if farm else "All Farms"])
    writer.writerow([])
    writer.writerow(["Disease Class", "Count", "Percentage"])
    for d in disease_stats:
        writer.writerow([d["name"], d["count"], f'{d["pct"]}%'])
    writer.writerow(["Total", total, "100%"])

    if not farm:
        disease_names = [d["name"] for d in disease_stats]
        writer.writerow([])
        writer.writerow(["Per-Farm Breakdown"])
        writer.writerow(["Farm ID", "Farm Name", "Owner", "Total Trees"] + disease_names)
        for f in Farm.objects.filter(owner=request.user).order_by("farm_id"):
            f_stats = {d["name"]: d["count"] for d in _get_full_disease_stats(request, f)}
            writer.writerow(
                [f.farm_id, f.name, f.owner_name, f.trees.count()]
                + [f_stats.get(name, 0) for name in disease_names]
            )

    return response


@login_required
def export_excel(request):
    # Exports a summary-only Excel file: overall disease totals and a per-farm breakdown.
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from django.http import HttpResponse

    farm = _get_farm_or_none(request)
    total, counts, pcts, diseased = _get_stats(request, farm)
    disease_stats = _get_full_disease_stats(request, farm)
    label = farm.farm_id if farm else "all_farms"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A2535", end_color="1A2535", fill_type="solid")
    title_font = Font(bold=True, size=13)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    ws.append(["RubberGuard Disease Detection Summary"])
    ws["A1"].font = title_font
    ws.append(["Scope", farm.name if farm else "All Farms"])
    ws.append([])

    ws.append(["Disease Class", "Count", "Percentage"])
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill

    for d in disease_stats:
        ws.append([d["name"], d["count"], f'{d["pct"]}%'])
    ws.append(["Total", total, "100%"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    if not farm:
        disease_names = [d["name"] for d in disease_stats]
        ws.append([])
        ws.append(["Per-Farm Breakdown"])
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=11)
        ws.append(["Farm ID", "Farm Name", "Owner", "Total Trees"] + disease_names)
        for cell in ws[ws.max_row]:
            cell.font = header_font
            cell.fill = header_fill
        for f in Farm.objects.filter(owner=request.user).order_by("farm_id"):
            f_stats = {d["name"]: d["count"] for d in _get_full_disease_stats(request, f)}
            ws.append(
                [f.farm_id, f.name, f.owner_name, f.trees.count()]
                + [f_stats.get(name, 0) for name in disease_names]
            )

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="rubberguard_summary_{label}.xlsx"'
    wb.save(response)
    return response


def _build_trend_chart(monthly, disease_stats, chart_width, chart_height):
    # Renders the monthly detection trend as a native ReportLab bar chart,
    # limited to the most recent 3 months. The previous version plotted the
    # entire scan history (15+ months) squeezed into one landscape-page
    # chart, which made individual bars unreadable. Only diseases that
    # actually occurred in this 3-month window get a legend entry, and the
    # legend now sits in its own horizontal row under the title instead of
    # a side column, so it's never competing with the bars for space.
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.legends import Legend
    from reportlab.lib import colors as rl_colors

    recent = monthly[-3:] if monthly else []
    months = [m["month"] for m in recent]
    series = [(d["name"], rl_colors.HexColor(d["color"])) for d in disease_stats
              if any(m["diseases"].get(d["name"], 0) for m in recent)]

    drawing = Drawing(chart_width, chart_height)
    drawing.add(String(chart_width / 2, chart_height - 14, "Monthly Detection Trend (last 3 months)",
                        fontSize=11, fontName="Helvetica-Bold", textAnchor="middle"))

    if not recent or not series:
        drawing.add(String(chart_width / 2, chart_height / 2, "No scan history in this window",
                            fontSize=9, textAnchor="middle"))
        return drawing

    trend_data = [[m["diseases"].get(name, 0) for m in recent] for name, _ in series]

    legend_y = chart_height - 34
    legend = Legend()
    legend.x = chart_width / 2
    legend.y = legend_y
    legend.dx = 8
    legend.dy = 8
    legend.fontSize = 8
    legend.alignment = "right"
    legend.boxAnchor = "n"
    legend.columnMaximum = 1
    legend.deltax = 14
    legend.colorNamePairs = [(color, name) for name, color in series]
    drawing.add(legend)

    bar = VerticalBarChart()
    bar.x = 55
    bar.y = 30
    bar.width = chart_width - 80
    bar.height = legend_y - 46
    bar.data = trend_data
    bar.categoryAxis.categoryNames = months
    bar.categoryAxis.labels.fontSize = 11
    bar.categoryAxis.labels.fontName = "Helvetica-Bold"
    bar.valueAxis.labels.fontSize = 9
    bar.valueAxis.valueMin = 0
    bar.valueAxis.forceZero = True
    bar.groupSpacing = 22
    bar.barSpacing = 3
    bar.barLabels.fontSize = 7.5
    bar.barLabelFormat = "%d"
    bar.barLabels.dy = 4
    for i, (_, color) in enumerate(series):
        bar.bars[i].fillColor = color
    drawing.add(bar)

    return drawing


def _build_donut(drawing, cx, cy, radius, data, colors_, cutout=0.65,
                  center_title=None, center_subtitle=None):
    # Draws a circular donut centered at (cx, cy): a Pie forced into a
    # perfect circle (equal width/height -- the old code used chart_width
    # and chart_height independently, which is what produced the "oblong"
    # egg-shaped pie) with a same-color-as-page hole punched in the middle
    # to fake a cutout, since ReportLab's Pie has no native donut mode.
    # cutout=0.65 matches the Chart.js `cutout: "65%"` used by the
    # doughnuts on the web Dashboard/Reports pages (dashboard.js /
    # reports.js) so the PDF and the web app look like the same design.
    from reportlab.graphics.shapes import String, Circle
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.lib import colors as rl_colors

    diameter = radius * 2
    pie = Pie()
    pie.x = cx - radius
    pie.y = cy - radius
    pie.width = diameter
    pie.height = diameter
    pie.data = data
    pie.labels = ["" for _ in data]
    pie.simpleLabels = 0
    pie.sideLabels = 0
    pie.slices.strokeWidth = 1.5
    pie.slices.strokeColor = rl_colors.white
    pie.slices.fontSize = 0
    for i, color in enumerate(colors_):
        pie.slices[i].fillColor = color
    drawing.add(pie)

    drawing.add(Circle(cx, cy, radius * cutout, fillColor=rl_colors.white, strokeColor=None))

    if center_title:
        drawing.add(String(cx, cy + (1 if center_subtitle else -4), center_title,
                            fontSize=15, fontName="Helvetica-Bold", textAnchor="middle",
                            fillColor=rl_colors.HexColor("#1a2535")))
    if center_subtitle:
        drawing.add(String(cx, cy - 13, center_subtitle,
                            fontSize=7.5, fontName="Helvetica", textAnchor="middle",
                            fillColor=rl_colors.HexColor("#6b7280")))


def _build_pie_chart(disease_stats, chart_width, chart_height, total=None):
    # Disease Distribution: a circular donut on the left (real cutout, not
    # an oval pie) plus a proper legend column on the right instead of the
    # old leader-line labels, which collided whenever one disease (usually
    # Healthy) dominates and the rest are thin slivers -- exactly what the
    # per-farm data here looks like.
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.graphics.charts.legends import Legend
    from reportlab.lib import colors as rl_colors

    nonzero = [d for d in disease_stats if d["count"] > 0]

    drawing = Drawing(chart_width, chart_height)
    drawing.add(String(chart_width / 2, chart_height - 14, "Disease Distribution",
                        fontSize=11, fontName="Helvetica-Bold", textAnchor="middle"))
    if not nonzero:
        drawing.add(String(chart_width / 2, chart_height / 2, "No data",
                            fontSize=9, textAnchor="middle"))
        return drawing

    radius = min(chart_width * 0.19, (chart_height - 30) / 2)
    cx = chart_width * 0.27
    cy = (chart_height - 16) / 2 + 4
    total = total if total else sum(d["count"] for d in nonzero)

    _build_donut(
        drawing, cx, cy, radius,
        data=[d["count"] for d in nonzero],
        colors_=[rl_colors.HexColor(d["color"]) for d in nonzero],
        center_title=f"{total:,}",
        center_subtitle="trees scanned",
    )

    legend = Legend()
    legend.x = chart_width * 0.52
    legend.y = chart_height - 34
    legend.dx = 9
    legend.dy = 9
    legend.fontSize = 8.5
    legend.alignment = "right"
    legend.columnMaximum = len(nonzero)
    legend.deltay = 15
    legend.colorNamePairs = [
        (rl_colors.HexColor(d["color"]), f'{d["name"]} — {d["count"]} ({d["pct"]}%)')
        for d in nonzero
    ]
    drawing.add(legend)
    return drawing


def _build_severity_pie(severity_counts, chart_width, chart_height, total=None):
    # Severity Distribution donut: same circular-cutout style as the disease
    # distribution chart above, using the exact tier colors from
    # SEVERITY_COLORS in reports.js. No leader-line labels or legend here on
    # purpose -- the Severity Distribution table already printed directly
    # above this chart carries the exact counts, so the donut's job is just
    # the at-a-glance shape, with the headline "% Healthy" in the center.
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.lib import colors as rl_colors

    tiers = [
        ("Healthy", severity_counts.get("Healthy", 0), rl_colors.HexColor("#28a745")),
        ("Mild", severity_counts.get("Mild", 0), rl_colors.HexColor("#fbbf24")),
        ("Moderate", severity_counts.get("Moderate", 0), rl_colors.HexColor("#f97316")),
        ("Severe", severity_counts.get("Severe", 0), rl_colors.HexColor("#dc2626")),
    ]
    nonzero = [(name, v, color) for name, v, color in tiers if v > 0]

    drawing = Drawing(chart_width, chart_height)
    drawing.add(String(chart_width / 2, chart_height - 14, "Severity Distribution",
                        fontSize=11, fontName="Helvetica-Bold", textAnchor="middle"))
    if not nonzero:
        drawing.add(String(chart_width / 2, chart_height / 2, "No data",
                            fontSize=9, textAnchor="middle"))
        return drawing

    total = total if total else sum(v for _, v, _ in nonzero)
    healthy_n = severity_counts.get("Healthy", 0)
    healthy_pct = round(healthy_n / total * 100, 1) if total else 0
    radius = min(chart_width * 0.32, (chart_height - 30) / 2)

    _build_donut(
        drawing, chart_width / 2, (chart_height - 16) / 2 + 4, radius,
        data=[v for _, v, _ in nonzero],
        colors_=[color for _, _, color in nonzero],
        center_title=f"{healthy_pct}%",
        center_subtitle="Healthy",
    )
    return drawing


def _build_most_affected_chart(most_affected, chart_width, chart_height):
    # Renders the "Most Affected Areas" farm ranking as a native ReportLab
    # horizontal bar chart. This panel exists on the web Reports page (as
    # HTML bar rows driven by _get_most_affected_farms) but was previously
    # missing from the PDF export entirely -- export_pdf() never called
    # _get_most_affected_farms(), so the data never reached the PDF at all.
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.graphics.charts.barcharts import HorizontalBarChart
    from reportlab.lib import colors as rl_colors

    drawing = Drawing(chart_width, chart_height)
    drawing.add(String(chart_width / 2, chart_height - 12, "Most Affected Areas (diseased tree count)",
                        fontSize=10, fontName="Helvetica-Bold", textAnchor="middle"))
    if most_affected:
        bar = HorizontalBarChart()
        bar.x = 110
        bar.y = 10
        bar.width = chart_width - 130
        bar.height = chart_height - 34
        bar.data = [[m["diseased"] for m in most_affected]]
        bar.categoryAxis.categoryNames = [m["label"] for m in most_affected]
        bar.categoryAxis.labels.fontSize = 7
        bar.valueAxis.labels.fontSize = 7
        bar.valueAxis.valueMin = 0
        bar.valueAxis.forceZero = True
        bar.bars[0].fillColor = rl_colors.HexColor("#dc2626")
        drawing.add(bar)
    else:
        drawing.add(String(chart_width / 2, chart_height / 2, "No farms registered yet",
                            fontSize=9, textAnchor="middle"))
    return drawing


@login_required
def export_pdf(request):
    # Exports a landscape PDF report with KPI summary, charts, and a per-tree
    # table. Landscape avoids column/text truncation that portrait caused
    # with this much side-by-side content.
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepInFrame
    from reportlab.lib.styles import getSampleStyleSheet
    from django.http import HttpResponse

    trees, label = _export_rows(request)
    farm = _get_farm_or_none(request)
    total, counts, pcts, diseased = _get_stats(request, farm)
    disease_stats = _get_full_disease_stats(request, farm)
    severity_counts = _get_severity_counts(request, farm)
    monthly = _get_monthly_trend(request, farm)
    effectiveness, _recent_ivs = _get_intervention_effectiveness(request, farm)
    disease_breakdown = _get_disease_breakdown(request, farm)
    confidence_summary = _get_confidence_summary(request, farm)
    most_affected = _get_most_affected_farms(request)
    key_insights = _get_key_insights(total, counts, pcts, diseased, effectiveness, disease_breakdown, confidence_summary)
    recommendations = _get_recommendations(disease_breakdown, effectiveness)
    generated_at = timezone.now()

    disease_names = [d["name"] for d in disease_stats]
    farm_summaries = []
    if not farm:
        for f in Farm.objects.filter(owner=request.user).order_by("farm_id"):
            f_stats = {d["name"]: d["count"] for d in _get_full_disease_stats(request, f)}
            farm_summaries.append(
                (f.farm_id, f.name, f.trees.count())
                + tuple(f_stats.get(name, 0) for name in disease_names)
            )

    intervention_qs = Intervention.objects.select_related("tree", "tree__farm").filter(tree__farm__owner=request.user)
    if farm:
        intervention_qs = intervention_qs.filter(tree__farm=farm)
    intervention_qs = intervention_qs.order_by("-date_performed")[:15]

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="rubberguard_report_{label}.pdf"'

    page_size = landscape(letter)
    page_width, page_height = page_size
    left_margin = right_margin = 0.5 * inch
    top_margin = bottom_margin = 0.5 * inch
    content_width = page_width - left_margin - right_margin

    doc = SimpleDocTemplate(
        response, pagesize=page_size,
        leftMargin=left_margin, rightMargin=right_margin,
        topMargin=top_margin, bottomMargin=bottom_margin,
    )
    styles = getSampleStyleSheet()
    elements = []

    title = farm.name if farm else "All Farms"
    elements.append(Paragraph("RubberGuard Disease Detection Report", styles["Title"]))
    elements.append(Paragraph(f"Scope: {title} &nbsp;&nbsp;|&nbsp;&nbsp; Generated: {timezone.localdate()}", styles["Normal"]))

    # Report metadata — which AI model produced the underlying classifications,
    # which report template version generated this file, and the exact
    # generation timestamp (not just the date), so any copy of this PDF can
    # be traced back to the data and detector version behind it.
    meta_style = styles["Normal"].clone("meta_style")
    meta_style.fontSize = 8
    meta_style.textColor = colors.HexColor("#6b7280")
    elements.append(Paragraph(
        f"AI Model: {AI_MODEL_VERSION} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"Report Version: {REPORT_VERSION} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"Generated at: {generated_at.strftime('%Y-%m-%d %H:%M %Z') or generated_at.strftime('%Y-%m-%d %H:%M')}",
        meta_style,
    ))
    elements.append(Spacer(1, 10))

    # Key Insights — plain-language, data-driven takeaways up front so a
    # non-technical reader gets the "so what" before the raw tables.
    elements.append(Paragraph("Key Insights", styles["Heading4"]))
    insight_style = styles["Normal"].clone("insight_style")
    insight_style.fontSize = 9
    insight_style.leftIndent = 10
    insight_style.spaceAfter = 3
    for point in key_insights:
        elements.append(Paragraph(f"•  {point}", insight_style))
    elements.append(Spacer(1, 12))

    # KPI summary row
    col_count = 1 + len(disease_stats)
    summary_data = [
        ["Total Trees"] + [d["name"] for d in disease_stats],
        [str(total)] + [str(d["count"]) for d in disease_stats],
    ]
    summary_table = Table(summary_data, colWidths=[content_width / col_count] * col_count)
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a2535")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 8))

    # Severity summary row
    severity_data = [
        ["Healthy", "Mild", "Moderate", "Severe"],
        [str(severity_counts["Healthy"]), str(severity_counts["Mild"]), str(severity_counts["Moderate"]), str(severity_counts["Severe"])],
    ]
    severity_table = Table(severity_data, colWidths=[content_width / 4] * 4)
    severity_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#dcfce7")),
        ("BACKGROUND", (1, 0), (1, -1), colors.HexColor("#fef3c7")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#fed7aa")),
        ("BACKGROUND", (3, 0), (3, -1), colors.HexColor("#fecaca")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(Paragraph("Severity Distribution", styles["Heading4"]))
    elements.append(severity_table)
    elements.append(Spacer(1, 8))

    # Severity donut — visual counterpart to the table above, matching the
    # "Disease Severity Distribution" chart on the web Reports page (same
    # tier colors) so this section isn't numbers-only in the PDF.
    sev_pie_h = 2.3 * inch
    sev_pie_w = 3.4 * inch
    sev_pie_drawing = _build_severity_pie(severity_counts, sev_pie_w, sev_pie_h, total=total)
    sev_pie_frame = KeepInFrame(sev_pie_w, sev_pie_h, [sev_pie_drawing])
    sev_pie_row = Table([[sev_pie_frame]], colWidths=[content_width])
    sev_pie_row.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))
    elements.append(sev_pie_row)
    elements.append(Spacer(1, 12))

    # Disease distribution pie — centered, full width available since the
    # trend bar chart (previously squeezed beside it) now gets its own
    # full-width row below with much more room to breathe.
    pie_h = 2.6 * inch
    pie_w = 5.6 * inch
    pie_drawing = _build_pie_chart(disease_stats, pie_w, pie_h, total=total)
    pie_frame = KeepInFrame(pie_w, pie_h, [pie_drawing])
    pie_row = Table([[pie_frame]], colWidths=[content_width])
    pie_row.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))
    elements.append(pie_row)
    elements.append(Spacer(1, 10))

    # Monthly detection trend — a grouped bar chart, given the full report
    # width and taller height so month labels and the 4-series legend
    # aren't cramped the way they were when sharing a row with the pie.
    trend_h = 2.6 * inch
    trend_drawing = _build_trend_chart(monthly, disease_stats, content_width, trend_h)
    trend_frame = KeepInFrame(content_width, trend_h, [trend_drawing])
    elements.append(trend_frame)
    elements.append(Spacer(1, 14))

    # Most Affected Areas — this panel exists on the web Reports page but
    # was previously dropped from the PDF export entirely: export_pdf()
    # never called _get_most_affected_farms(). Added here for parity.
    if most_affected:
        elements.append(Paragraph("Most Affected Areas", styles["Heading4"]))
        affected_h = 2.2 * inch
        affected_drawing = _build_most_affected_chart(most_affected, content_width, affected_h)
        affected_frame = KeepInFrame(content_width, affected_h, [affected_drawing])
        elements.append(affected_frame)
        elements.append(Spacer(1, 14))

    # Per-farm breakdown (only when viewing all farms)
    if farm_summaries:
        farm_rows = [["Farm ID", "Farm Name", "Total"] + disease_names] + \
            [[str(v) for v in row] for row in farm_summaries]
        fixed_frac = 0.12 + 0.28 + 0.12  # Farm ID, Farm Name, Total
        disease_frac = (1 - fixed_frac) / max(len(disease_names), 1)
        col_fracs = [0.12, 0.28, 0.12] + [disease_frac] * len(disease_names)
        farm_table = Table(farm_rows, colWidths=[content_width * f for f in col_fracs])
        farm_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(Paragraph("Per-Farm Breakdown", styles["Heading4"]))
        elements.append(farm_table)
        elements.append(Spacer(1, 14))

    # Recommendations — short, actionable next steps derived from the
    # disease breakdown and intervention track record above.
    elements.append(Paragraph("Recommendations", styles["Heading4"]))
    rec_style = styles["Normal"].clone("rec_style")
    rec_style.fontSize = 9
    rec_style.leftIndent = 10
    rec_style.spaceAfter = 3
    for rec in recommendations:
        elements.append(Paragraph(f"•  {rec}", rec_style))
    elements.append(Spacer(1, 14))

    # Intervention Effectiveness — what % of trees treated with each action
    # are currently Healthy, i.e. how well that intervention has worked.
    if effectiveness:
        eff_rows = [["Intervention", "Trees Treated", "Now Healthy", "Recovery Rate"]] + [
            [e["action"], str(e["treated"]), str(e["recovered"]), f"{e['recovery_pct']}%"]
            for e in effectiveness
        ]
        eff_table = Table(eff_rows, colWidths=[content_width * f for f in [0.4, 0.2, 0.2, 0.2]])
        eff_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a2535")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(Paragraph("Intervention Effectiveness", styles["Heading4"]))
        elements.append(eff_table)
        elements.append(Spacer(1, 14))

    # Per-tree table — capped to a page-friendly sample since a full
    # multi-thousand-row dump makes the PDF slow to generate and awkward
    # to read. CSV/Excel exports remain uncapped for the complete dataset.
    PDF_TREE_ROW_LIMIT = 75
    tree_list = list(trees[:PDF_TREE_ROW_LIMIT])
    total_tree_count = trees.count()

    heading = "Tree-Level Detail"
    if total_tree_count > PDF_TREE_ROW_LIMIT:
        heading += f" (showing {PDF_TREE_ROW_LIMIT} of {total_tree_count} — use CSV/Excel export for the full list)"
    elements.append(Paragraph(heading, styles["Heading4"]))

    tree_rows = [["Tree ID", "Farm", "Disease", "Conf. %", "Date Scanned"]]
    for t in tree_list:
        tree_rows.append([t.tree_id, t.farm.farm_id, t.disease, f"{t.confidence}%", str(t.date_scanned)])

    col_fractions = [0.16, 0.18, 0.32, 0.16, 0.18]
    tree_table = Table(tree_rows, colWidths=[content_width * f for f in col_fractions], repeatRows=1)
    tree_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(tree_table)
    elements.append(Spacer(1, 14))

    # Recent interventions
    if intervention_qs:
        elements.append(Paragraph("Recent Interventions", styles["Heading4"]))
        iv_rows = [["Tree ID", "Farm", "Action", "Date", "Notes"]]
        for iv in intervention_qs:
            iv_rows.append([iv.tree.tree_id, iv.tree.farm.farm_id, iv.action, str(iv.date_performed), (iv.notes or "—")[:40]])
        iv_table = Table(iv_rows, colWidths=[content_width * f for f in [0.12, 0.12, 0.22, 0.12, 0.42]])
        iv_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f0fdf4")),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ALIGN", (1, 0), (3, -1), "CENTER"),
            ("ALIGN", (4, 0), (4, -1), "LEFT"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(iv_table)

    doc.build(elements)
    return response


def csrf_failure(request, reason=""):
    # A CSRF mismatch on /login/ is usually a dropped mobile connection that
    # prevented the csrftoken cookie from being set before the form was
    # submitted (see server error log: "OSError: write error" immediately
    # preceding these). Instead of showing Django's raw 403 page, send the
    # user back to a fresh login form with a friendly explanation.
    messages.error(request, "Your session expired or the connection dropped. Please log in again.")
    return redirect("login")


@staff_member_required
def admin_dashboard(request):
    # A single operational dashboard: add disease classes without leaving
    # the app's own UI (full edit of an existing one still goes through
    # /admin/ for all the fields), plus storage and AI-model status at a
    # glance. Restricted to staff accounts (create one via
    # `manage.py createsuperuser` or the /admin/ Users page).
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        if not name:
            messages.error(request, "Disease name is required.")
        elif DiseaseClass.objects.filter(name__iexact=name).exists():
            messages.error(request, f'"{name}" already exists.')
        else:
            DiseaseClass.objects.create(
                name=name,
                description=request.POST.get("description", "").strip(),
                display_order=DiseaseClass.objects.count(),
                color_hex=request.POST.get("color_hex", "#6c757d").strip(),
                marker_key=request.POST.get("marker_key", "").strip() or name.lower().replace(" ", "_"),
                danger_rank=int(request.POST.get("danger_rank") or 1),
                is_healthy=request.POST.get("is_healthy") == "on",
                recommendation_mild=request.POST.get("recommendation_mild", "").strip(),
                recommendation_moderate=request.POST.get("recommendation_moderate", "").strip(),
                recommendation_severe=request.POST.get("recommendation_severe", "").strip(),
            )
            messages.success(request, f'Added disease class "{name}".')
        return redirect("admin_dashboard")

    diseases = DiseaseClass.objects.all().order_by("display_order", "name")

    # AI model analytics are necessarily a stub until real per-inference
    # logging exists -- there's no trained model wired in yet for most
    # deployments of this app. What's shown is what can be known today:
    # whether an endpoint is configured, and aggregate counts as a rough
    # proxy for "how much scanning has happened."
    ai_status = {
        "enabled": ai_inference.AI_MODEL_ENABLED,
        "endpoint_host": (
            ai_inference.AI_MODEL_ENDPOINT_URL.split("//")[-1].split("/")[0]
            if ai_inference.AI_MODEL_ENABLED else None
        ),
        "total_scans": ScanHistory.objects.count(),
        "total_trees": RubberTree.objects.count(),
        "avg_confidence": round(RubberTree.objects.aggregate(avg=Avg("confidence"))["avg"] or 0, 1),
    }

    ctx = _base_context(request, _get_farm_or_none(request))
    ctx.update({
        "page": "admin_dashboard",
        "diseases": diseases,
        "storage_summary": storage_stats.get_storage_summary(),
        "ai_status": ai_status,
    })
    return render(request, "admin_dashboard.html", ctx)
